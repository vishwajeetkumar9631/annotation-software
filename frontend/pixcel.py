import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from PIL import Image


def convert_image_to_colored_excel(
    image_path, output_excel_path, max_dimension=100
):
    # 1. Verify if the file exists
    if not os.path.exists(image_path):
        print(f"Error: The file at {image_path} was not found.")
        return

    print("Opening and processing image...")

    # 2. Load the image and force it to RGB mode
    img = Image.open(image_path)
    img = img.convert("RGB")

    # 3. Downscale the image to prevent Excel from lagging
    # Setting this to 100px ensures it fits safely in standard Excel limits
    img.thumbnail((max_dimension, max_dimension))

    pixels = img.load()
    width, height = img.size

    # 4. Create an Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pixel Art"

    print(
        f"Painting {width}x{height} cells ({width * height} total cells) in Excel..."
    )

    # 5. Loop through every pixel and paint the matching cell
    for y in range(height):
        # Optional progress display for larger renders
        if y % 20 == 0:
            print(f"Processing row {y}/{height}...")

        # Set row height to make it square (default width is usually around 8-10)
        ws.row_dimensions[y + 1].height = 15

        for x in range(width):
            r, g, b = pixels[x, y]

            # Convert RGB values to hex string (e.g., "FFAABB")
            hex_color = f"{r:02X}{g:02X}{b:02X}"

            # Grab the specific cell (Excel uses 1-based indexing)
            cell = ws.cell(row=y + 1, column=x + 1)

            # Apply the solid background color
            cell.fill = PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )

    # Set column widths to make them square grids
    for col in range(1, width + 1):
        # Openpyxl dimensions aren't 1:1, ~2.5 creates a decent square
        ws.column_dimensions[get_column_letter(col)].width = 2.5

    # 6. Save the actual .xlsx file
    print(f"Saving file to {output_excel_path}...")
    wb.save(output_excel_path)
    print("Done! Open the saved file in Excel to see your image!")


# Execution Parameters
input_file = r"C:\Users\ankit\Downloads\WhatsApp Image 2026-04-21 at 4.44.41 PM (2).jpeg"
output_file = r"C:\Users\ankit\Downloads\pixel_PCBportrait.xlsx"

# Trigger execution
convert_image_to_colored_excel(input_file, output_file)
